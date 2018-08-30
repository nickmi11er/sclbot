# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from datetime import datetime, timedelta
import const
import date_manager
import re
import subprocess
from models.user import User

pt = re.compile(r'(?:\s*([0-9]+(?:,\s*[0-9]+)*)+\s*(–Ω|—á)+\s*)?\s*(.+)', re.UNICODE)
wb2 = load_workbook(filename=const.assets_dir + '/scl.xlsx')
ws2 = wb2.worksheets[0]

start_dt = datetime.strptime('03.09.2018', '%d.%m.%Y')
end_dt = datetime.strptime('31.05.2019', '%d.%m.%Y')
start_holy_dt = datetime.strptime('29.05.2019', '%d.%m.%Y')

scl_time = {
    1: "09:00-10:30",
    2: "10:40-12:10",
    3: "13:00-14:30",
    4: "14:40-16:10",
    5: "16:20-17:50",
    6: "18:00-19:30"
}

def updscl():
    subprocess.call('sh ../assets/dlscl.sh', shell=True)
    global wb2
    global ws2
    wb2 = load_workbook(filename=const.assets_dir + '/scl.xlsx')
    ws2 = wb2.worksheets[0]


# –í–æ–∑–≤—Ä–∞—â–∞–µ—Ç —Å–ª–µ–¥—É—é—â–∏–π —Å–∏–º–≤–æ–ª –≤ –∞–ª—Ñ–∞–≤–∏—Ç–µ –¥–ª—è –∏–Ω–∫—Ä–µ–º–µ–Ω—Ç—ã —Å—Ç–æ–ª–±—Ü–∞ —Ç–∞–±–ª–∏—Ü—ã
# TODO –¥–æ–±–∞–≤–∏—Ç—å –ª–æ–≥–∏–∫—É –ø—Ä–∏ –ø—Ä–∏–≤—ã—à–µ–Ω–∏–∏ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –±—É–∫–≤ –≤ –∞–ª—Ñ–∞–≤–∏—Ç–µ
def inc_col_name(col_name):
    if col_name == 'Z':
        return 'AA'
    elif len(col_name) > 1:
        return 'A' + inc_col_name(col_name[1])
    else:
        return chr(ord(col_name) + 1)


# –†–∞–∑–±–∏–≤–∞–µ—Ç —Å—Ç—Ä–æ–∫—É –Ω–∞ —Ç—Ä–∏ —Å–æ—Å—Ç–∞–≤–ª—è—é—â–∏–µ: 
# 1 –≥—Ä—É–ø–ø–∞ - –Ω–æ–º–µ—Ä–∞ –Ω–µ–¥–µ–ª—å, –≤ –∫–æ—Ç–æ—Ä—ã—Ö —Å—Ç–æ–∏—Ç –ø—Ä–µ–¥–º–µ—Ç 
# 3 –≥—Ä—É–ø–ø–∞ - –Ω–∞–∑–≤–∞–Ω–∏–µ —Å–∞–º–æ–≥–æ –ø—Ä–µ–¥–º–µ—Ç–∞ 
# –ù–∞ –æ—Å–Ω–æ–≤–∞–Ω–∏–∏ –ø–µ—Ä–µ–¥–∞–Ω–Ω–æ–≥–æ –Ω–æ–º–µ—Ä–∞ –Ω–µ–¥–µ–ª–∏, –≤–æ–∑–≤—Ä–∞—â–∞–µ—Ç –Ω–∞–∑–≤–∞–Ω–∏–µ –ø—Ä–µ–¥–º–µ—Ç–∞
def choose_task(entity, wleft):
    entity = entity.encode('utf-8')
    res = pt.match(entity)
    if res is not None and res.group(1) is not None:
        num_weeks = re.split(r',', res.group(1).replace(' ', ''))
        for num in num_weeks:
            if int(num) == int(wleft):
                return res.group(3)
    else:
        return 'kostil'
            

def _get_scl(gp_nm, date):
    date = date.replace(hour=0, minute=0, second=0, microsecond=0)
    result = []
    loaded_tasks = 0
    wleft = ((date - start_dt).days + 1) / 7 + 1    # –Ω–æ–º–µ—Ä —Ç–µ–∫—É—â–µ–π –Ω–µ–¥–µ–ª–∏

    wdoffset = 3    # –æ—Ç—Å—Ç—É–ø –ø–æ —è—á–µ–π–∫–∞–º
    # –µ—Å–ª–∏ –∑–∞–ø—Ä–∞—à–∏–≤–∞–µ–º—ã–π –¥–µ–Ω—å –Ω–µ –ø–æ–Ω–µ–¥–µ–ª—å–Ω–∏–∫, –¥–æ–±–∞–≤–ª—è–µ–º —Å–º–µ—â–µ–Ω–∏–µ
    if date.weekday() != 0:
        wdoffset = wdoffset + date.weekday() * 12

    findall = False
    # –ø–µ—Ä–µ–±–∏—Ä–∞–µ–º —É—á–µ–±–Ω—ã–µ –≥—Ä—É–ø–ø—ã 
    for col in ws2.iter_cols(min_col=1):
        if findall:
            break
        for cell in col:
            cv = cell.value
            if cv and int(cell.row) == 2:
                if isinstance(cv, basestring):
                    cv = cv.replace(' ', '')
            # –Ω–∞—à–ª–∏ –Ω—É–∂–Ω—É—é –≥—Ä—É–ø–ø—É
            if cv == gp_nm:
                cc = cell.column
                # –ø–∞—Ä—Å–∏–º —Ä–∞—Å–ø–∏—Å–∞–Ω–∏–µ –Ω–∞ –∑–∞–ø—Ä–æ—à–µ–Ω–Ω—ã–π –¥–µ–Ω—å
                # –ø–µ—Ä–µ–±–∏—Ä–∞–µ–º —Å–ª–µ–¥—É—é—â–∏–µ 12 —è—á–µ–µ–∫ —Å —à–∞–≥–æ–º 2 —Å —É—á–µ—Ç–æ–º —á–µ—Ç–Ω–æ—Å—Ç–∏ –Ω–µ–¥–µ–ª–∏
                for i in range(wdoffset + 1, wdoffset + 13, 2):
                    loaded_tasks += 1
                    daynum = i
                    if wleft % 2 == 0:
                        daynum = daynum + 1

                    val = ws2[cc + str(daynum)].value
                    if val is None:
                        continue
                    
                    # –ø–∞—Ä—Å–∏–º –≤—Ä–µ–º—è –ø–∞—Ä—ã, —Ç–∏–ø –∑–∞–Ω—è—Ç–∏–π –∏ –∞—É–¥–∏—Ç–æ—Ä–∏—é
                    time = scl_time[loaded_tasks]

                    cn = inc_col_name(cell.column)
                    fullnm = cn + str(daynum)
                    task_tp = ws2[fullnm].value
                    # –µ—Å–ª–∏ —Ç–∏–ø –∑–∞–Ω—è—Ç–∏—è - –ø—Ä–∞–∫—Ç–∏–∫–∞, –¥–æ–±–∞–≤–ª—è–µ–º —Å–∏–º–≤–æ–ª –≤–æ—Å–∫–ª–∏—Ü–∞—Ç–µ–ª—å–Ω–æ–≥–æ –∑–Ω–∞–∫–∞
                    if task_tp:
                        if task_tp.encode('utf-8') == '–ø—Ä':
                            task_tp = ' ‚ùó'
                        elif task_tp.encode('utf-8') == '–ª—Ä':
                            task_tp = ' üë©‚Äçüî¨'
                        else:
                            task_tp = ''
                    else:
                        task_tp = ''

                    cn = inc_col_name(cn)
                    cn = inc_col_name(cn)
                    fullnm = cn + str(daynum)
                    cl_num = ws2[fullnm].value
                    if cl_num:
                        if isinstance(cl_num, long) or isinstance(cl_num, float):
                            cl_num = str(int(cl_num))
                        else:
                            cl_num = cl_num.encode('utf-8')
                        cl_num = '(' + cl_num + ')'
                    else:
                        cl_num = ''

                    # –ø—Ä–æ–≤–µ—Ä—è–µ–º —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –ª–∏ –≤–∞—Ä–∏–∞—Ç–∏–≤–Ω–æ—Å—Ç—å –ø–∞—Ä –Ω–∞ –æ–¥–∏–Ω –∏ —Ç–æ—Ç –∂–µ –ø—Ä–æ–º–µ–∂—É—Ç–æ–∫ –≤—Ä–µ–º–µ–Ω–∏
                    ent = re.split(r'\n', val)
                    if len(ent) == 1:
                        res = choose_task(ent[0], wleft)
                        if res is not None:
                            if res == 'kostil':
                                res = ent[0].encode('utf-8')
                            result.append('{} {} {} {}'.format(time, res, cl_num, task_tp))
                    # –µ—Å–ª–∏ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –≤–∞—Ä–∏–∞—Ç–∏–≤–Ω–æ—Å—Ç—å, –≤—ã–±–∏—Ä–∞–µ–º –Ω–µ–æ–±—Ö–æ–¥–∏–º—ã–π –ø—Ä–µ–¥–º–µ—Ç
                    elif len(ent) > 1:
                        ch_flag = 0
                        for i in ent:
                            res = choose_task(i, wleft)
                            if res is not None:
                                if res == 'kostil':
                                    continue
                                cl_nums = re.split(r'\n', cl_num)
                                if len(cl_nums) == 2:
                                    cl_num = cl_nums[ch_flag]
                                result.append('{} {} {} {}'.format(time, res, cl_num, task_tp))
                                break
                            ch_flag += 1
                findall = True
                break
    return result     
        


def get_scl_with(dt, id):
    date = datetime.now()
    user = User.get(id)
    if dt is not None:
        date = dt

    if date < start_dt:
        return u'–£—á–µ–±–∞ –µ—â–µ –Ω–µ –Ω–∞—á–∞–ª–∞—Å—å'

    if date >= end_dt:
        return u'–£—á–µ–±–∞ –∑–∞–∫–æ–Ω—á–∏–ª–∞—Å—å. –£–¥–∞—á–∏ –Ω–∞ —Å–µ—Å—Å–∏–∏'

    if date >= start_holy_dt:
        return u'–õ–µ—Ç–Ω–∏–µ –∫–∞–Ω–∏–∫—É–ª—ã. –û—Ç–¥—ã—Ö–∞–π!'

    out = '–†–∞—Å–ø–∏—Å–∞–Ω–∏–µ –ø–∞—Ä –Ω–∞ {} ({}): \n\n'.format(date.strftime('%d.%m.%Y'), date_manager.rus_week_day[date.weekday()])
    res = _get_scl(user.group_name, date)

    if res:
        for r in res:
            out = out + r + "\n"
    else:
        out = out + "–ü–∞—Ä –Ω–µ—Ç. –û—Ç–¥—ã—Ö–∞–π!\n"

    return out



def get_week_scl(dt, id):
    date = datetime.now()
    user = User.get(id)
    if dt:
        date = dt      

    start = date - timedelta(days=date.weekday())
    end = start + timedelta(days=6)  

    if start < start_dt:
        return u'–£—á–µ–±–∞ –µ—â–µ –Ω–∞ –Ω–∞—á–∞–ª–∞—Å—å'  

    out = ''
    for i in range(0, 6):
        if i > 0:
            out = out + '\n'
        out = out + get_scl_with(date_manager.get_day_over(i, start), id)
        
    return out
