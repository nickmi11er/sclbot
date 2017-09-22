# coding=utf-8
import logging
from openpyxl import load_workbook
from datetime import datetime

logging.basicConfig(filename='log.txt', level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

wb = load_workbook(filename='schedule.xlsx')
ws = wb.worksheets[0]

scl_time = {
    1: "09:00-10:30",
    2: "10:40-12:10",
    3: "13:00-14:30",
    4: "14:40-16:10",
    5: "16:20-17:50",
    6: "18:00-19:30"
}


# Возвращает следующий символ в алфавите для инкременты столбца таблицы
# TODO добавить логику при привышении количества букв в алфавите
def inc_col_name(col_name):
    return chr(ord(col_name) + 1)


def schedule_time(num):
    return scl_time[num]


def get_with(time):
    time = time.replace(hour=0, minute=0, second=0, microsecond=0)
    result = []
    loaded_cells = 0
    date_is_find = False

    for column in ws.iter_cols(min_col=1):

        for cell in column:

            if cell.value:

                if not date_is_find:
                    if isinstance(cell.value, datetime):
                        if time == cell.value:
                            date_is_find = True
                            continue

                if date_is_find and loaded_cells < 6:
                    loaded_cells += 1

                    if cell.value == "-":
                        continue

                    task_name = schedule_time(loaded_cells) + "     " + cell.value

                    # Если есть номер аудитории, добавляем в строку
                    cell_name = inc_col_name(cell.column) + str(cell.row)

                    if ws[cell_name].value:
                        val = ws[cell_name].value
                        if isinstance(val, float):
                            val = str(val)

                        task_name += "  " + val

                    result.append(task_name)

    return result


def get_academy_plan(cursor):
    res = ""

    exams = cursor.execute("SELECT name FROM subjects s WHERE s.exam = 'true'").fetchall()
    res += u"Экзамены: \n"
    i = 1
    for exam in exams:
        res += str(i) + ": " + exam[0] + '\n'
        i += 1

    res += u"Всего экзаменов: " + str(i - 1) + '\n'

    res += '\n'
    credits = cursor.execute("SELECT name FROM subjects s WHERE s.exam = 'false'").fetchall()
    res += u"Зачеты: \n"
    i = 1
    for credit in credits:
        res += str(i) + ": " + credit[0] + '\n'
        i += 1

    res += u"Всего зачетов: " + str(i - 1) + '\n'

    return res
